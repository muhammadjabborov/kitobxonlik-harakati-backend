import openpyxl

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.common.models import District, Region, School


class Command(BaseCommand):
    help = "Load schools from school_data/*.xlsx (all_regions_schools.xlsx + tashkent_schools.xlsx)"

    TASHKENT_REGION_NAME = "Toshkent shahri"

    def add_arguments(self, parser):
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete all existing schools before loading.",
        )

    def handle(self, *args, **options):
        if options["clear"]:
            deleted, _ = School.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"Deleted {deleted} existing School rows."))

        base = settings.BASE_DIR / "excel_data"

        total_created = 0
        total_skipped = 0
        total_missing_district = 0

        with transaction.atomic():
            created, skipped, missing = self._load_all_regions(base / "all_regions_schools.xlsx")
            total_created += created
            total_skipped += skipped
            total_missing_district += missing

            created, skipped, missing = self._load_tashkent(base / "tashkent_schools.xlsx")
            total_created += created
            total_skipped += skipped
            total_missing_district += missing

        self.stdout.write(
            self.style.SUCCESS(
                f"Done! Created: {total_created} schools, "
                f"Skipped (already existed): {total_skipped}, "
                f"Missing district (auto-created): {total_missing_district}."
            )
        )

    def _clean(self, value):
        if value is None:
            return None
        return str(value).strip()

    def _load_all_regions(self, path):
        self.stdout.write(f"Loading {path.name}...")
        wb = openpyxl.load_workbook(filename=path, read_only=True)
        ws = wb.active

        region_cache: dict[str, Region] = {}
        district_cache: dict[tuple, District] = {}
        created = skipped = missing_district = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 5:
                continue
            region_name = self._clean(row[1])
            district_name = self._clean(row[2])
            school_name = self._clean(row[3])
            if not region_name or not district_name or not school_name:
                continue

            region = self._get_region(region_name, region_cache)
            district, district_was_created = self._get_district(region, district_name, district_cache)
            if district_was_created:
                missing_district += 1

            _, was_created = School.objects.get_or_create(district=district, name=school_name)
            if was_created:
                created += 1
            else:
                skipped += 1

        self.stdout.write(f"  → {created} created, {skipped} already existed.")
        return created, skipped, missing_district

    def _load_tashkent(self, path):
        if not path.exists():
            self.stdout.write(self.style.WARNING(f"  Skipping {path.name} — file not found."))
            return 0, 0, 0

        self.stdout.write(f"Loading {path.name}...")
        wb = openpyxl.load_workbook(filename=path, read_only=True)
        ws = wb.active

        region_cache: dict[str, Region] = {}
        district_cache: dict[tuple, District] = {}
        created = skipped = missing_district = 0

        region = self._get_region(self.TASHKENT_REGION_NAME, region_cache)

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 4:
                continue
            district_name = self._clean(row[1])
            school_name = self._clean(row[2])
            if not district_name or not school_name:
                continue

            district, district_was_created = self._get_district(region, district_name, district_cache)
            if district_was_created:
                missing_district += 1

            _, was_created = School.objects.get_or_create(district=district, name=school_name)
            if was_created:
                created += 1
            else:
                skipped += 1

        self.stdout.write(f"  → {created} created, {skipped} already existed.")
        return created, skipped, missing_district

    def _get_region(self, name, cache):
        if name in cache:
            return cache[name]
        region, _ = Region.objects.get_or_create(name=name)
        cache[name] = region
        return region

    def _get_district(self, region, name, cache):
        key = (region.pk, name)
        if key in cache:
            return cache[key], False
        district, was_created = District.objects.get_or_create(region=region, name=name)
        cache[key] = district
        return district, was_created
